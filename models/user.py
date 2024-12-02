import openpyxl

# Функция для регистрации пользователя
def register_user(user_id, first_name, role):
    try:
        workbook = openpyxl.load_workbook('users.xlsx')
        sheet = workbook.active
        # Проверяем, зарегистрирован ли уже пользователь
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return False  # Пользователь уже зарегистрирован
        # Добавляем нового пользователя
        sheet.append([user_id, first_name, role])
        workbook.save('users.xlsx')
        return True
    except Exception as e:
        print(f"Error registering user: {e}")
        return False

def get_user_role(user_id):
    try:
        workbook = openpyxl.load_workbook('users.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return row[4]  # Возвращаем роль пользователя (в столбце 5)
        return None
    except Exception as e:
        print(f"Error getting user role: {e}")
        return None

def add_user_to_excel(user_id, first_name, last_name, group, role):
    try:
        workbook = openpyxl.load_workbook('users.xlsx')
        sheet = workbook.active
        sheet.append([user_id, first_name, last_name, group, role])
        workbook.save('users.xlsx')
    except Exception as e:
        print(f"Error adding user to Excel: {e}")