import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

FILE_NAME = 'project_manager.xlsx'

# Инициализация нового Excel файла
def initialize_excel():
    if not os.path.exists(FILE_NAME):
        workbook = Workbook()

        # Лист для пользователей
        sheet_users = workbook.active
        sheet_users.title = "Users"
        sheet_users.append(["Telegram ID", "Username", "Role"])  # Заголовки

        # Применяем стили
        for cell in sheet_users[1]:
            cell.font = Font(name='Arial', size=12, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Лист для проектов
        sheet_projects = workbook.create_sheet("Projects")
        sheet_projects.append([
            "Project ID", "Student ID", "Student Username", "Teacher ID", "Title", "Description", "Status", "Date Created", "Date Updated"
        ])  # Заголовки

        # Применяем стили
        for cell in sheet_projects[1]:
            cell.font = Font(name='Arial', size=12, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Лист для сообщений
        sheet_messages = workbook.create_sheet("Messages")
        sheet_messages.append(["Message ID", "Sender ID", "Receiver ID", "Message Text", "Date Sent", "Status", "Project ID"])

        # Применяем стили
        for cell in sheet_messages[1]:
            cell.font = Font(name='Arial', size=12, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(FILE_NAME)


def add_user_to_excel(telegram_id, username, role):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook["Users"]

    # Проверяем, зарегистрирован ли пользователь
    for row in sheet.iter_rows(values_only=True):
        if row[0] == telegram_id:
            return  # Пользователь уже существует, ничего не делаем

    # Добавляем нового пользователя
    sheet.append([telegram_id, username, role])
    workbook.save(FILE_NAME)


def add_project_to_excel(student_id, username, teacher_id, title, description, status):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook["Projects"]

    # Определяем новый ID для проекта
    project_id = len(sheet['A']) + 1  # Или можно найти первую пустую строку

    # Добавляем новый проект
    date_created = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([project_id, student_id, username, teacher_id, title, description, status, date_created, None])

    workbook.save(FILE_NAME)


def get_projects_by_student(student_id):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook["Projects"]

    projects = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == student_id:
            projects.append({
                'id': row[0],
                'student_id': row[1],
                'student_username': row[2],
                'teacher_id': row[3],
                'title': row[4],
                'description': row[5],
                'status': row[6],
                'date_created': row[7],
                'date_updated': row[8]
            })
    return projects


def get_projects_by_teacher(teacher_id):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook["Projects"]

    projects = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] == teacher_id:
            projects.append({
                'id': row[0],
                'student_id': row[1],
                'student_username': row[2],
                'teacher_id': row[3],
                'title': row[4],
                'description': row[5],
                'status': row[6],
                'date_created': row[7],
                'date_updated': row[8]
            })
    return projects


def add_message_to_excel(sender_id, receiver_id, message_text, project_id):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook["Messages"]

    # Определяем новый ID для сообщения
    message_id = len(sheet['A']) + 1  # Аналогично, проверяем количество строк

    # Добавляем новое сообщение
    date_sent = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([message_id, sender_id, receiver_id, message_text, date_sent, 'unread', project_id])

    workbook.save(FILE_NAME)


def get_new_messages_for_user(user_id):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook["Messages"]

    new_messages = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == user_id and row[5] == 'unread':
            new_messages.append({
                'message_id': row[0],
                'sender_id': row[1],
                'receiver_id': row[2],
                'message_text': row[3],
                'date_sent': row[4],
                'project_id': row[6]
            })

    return new_messages


def mark_message_as_read(message_id):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook["Messages"]

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == message_id:
            row[5].value = 'read'  # Помечаем сообщение как прочитанное
            break

    workbook.save(FILE_NAME)


# Функция для получения всех проектов из Excel
def get_all_projects():
    """Получить все проекты из Excel"""
    try:
        df = pd.read_excel(FILE_NAME, sheet_name="Projects")  # Чтение Excel в DataFrame
        # Преобразуем DataFrame в список словарей с явным указанием колонок
        projects = df.to_dict(orient='records')
        for project in projects:
            # Убедимся, что все обязательные ключи присутствуют
            if 'student_username' not in project:
                project['student_username'] = 'Не указан'
            if 'student_id' not in project:
                project['student_id'] = 'Не указан'
        return projects
    except Exception as e:
        print(f"Ошибка при чтении Excel файла: {e}")
        return []

