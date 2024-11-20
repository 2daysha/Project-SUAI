from telebot import types
from student_handlers import student_menu, project_work_menu, my_projects_command, propose_project_topic, get_user_role
from excel_manager import get_projects_by_teacher, add_project_to_excel, add_user_to_excel, FILE_NAME, get_all_projects, get_new_messages_for_user, mark_message_as_read
from message_sender import send_message_to_student, check_for_new_messages
from teacher_handlers import teacher_menu
from openpyxl import load_workbook
import openpyxl

# ID преподавателя
TEACHER_ID = 1008919333

def view_projects_command(bot, message):
    try:
        # Получаем проекты преподавателя
        projects = get_projects_by_teacher(message.chat.id)

        if not projects:
            bot.send_message(message.chat.id, "У вас нет проектов.")
            return

        # Отображаем проекты
        for project in projects:
            # Отладочная информация для диагностики
            print(f"Проект: {project}")  # Выводим проект для отладки

            # Пытаемся извлечь значения
            student_username = project.get('student_username', 'Не указан')
            student_id = project.get('student_id', 'Не указан')

            bot.send_message(
                message.chat.id,
                f"Проект: {project['title']}\n"
                f"Описание: {project['description']}\n"
                f"Студент: {student_username} (ID: {student_id})\n"
                f"Статус: {project['status']}"
            )
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка при получении проектов: {str(e)}")


def register_handlers(bot):
    @bot.message_handler(commands=['start'])
    def start(message):
        """Начальная команда для регистрации или отображения меню."""
        user_id = message.chat.id
        username = message.from_user.username
        role = get_user_role(user_id)

        if not role:
            # Если пользователь не зарегистрирован, предлагаем выбрать роль
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add("Зарегистрироваться как студент", "Зарегистрироваться как преподаватель")
            bot.send_message(
                user_id,
                f"Привет, {username}! Ваш Telegram ID: {user_id}. Выберите вашу роль:",
                reply_markup=markup
            )
        else:
            # Показать меню в зависимости от роли
            show_menu(bot, message, role)

    def get_all_students():
        """Получение списка всех студентов из Excel."""
        workbook = openpyxl.load_workbook(FILE_NAME)
        sheet = workbook["Users"]

        students = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == "student":  # Проверяем, что это студент
                students.append({
                    'student_id': row[0],  # Telegram ID студента
                    'username': row[1]
                })
        return students
    def show_menu(bot, message, role):
        """Отображение меню в зависимости от роли пользователя."""
        if role == "student":
            student_menu(bot, message)
        elif role == "teacher":
            teacher_menu(bot, message)

    @bot.message_handler(commands=['send_message_to_students'])
    def send_message_to_students(message):
        """Отправка сообщения всем студентам."""
        # Получаем список студентов
        students = get_all_students()  # Функция, которая получает список студентов из Excel или базы данных

        # Сообщение, которое преподаватель хочет отправить
        bot.send_message(message.chat.id, "Введите текст сообщения для студентов:")

        # Регистрация следующего шага для получения сообщения
        bot.register_next_step_handler(message, lambda msg: send_message_to_all_students(msg, students))

    def send_message_to_all_students(message, students):
        """Отправка сообщения всем студентам."""
        message_text = message.text  # Текст сообщения от преподавателя

        for student in students:
            student_id = student['student_id']  # Получаем ID студента
            bot.send_message(student_id, message_text)  # Отправляем сообщение студенту

        bot.send_message(message.chat.id, "Сообщение отправлено всем студентам.")

    @bot.message_handler(func=lambda msg: msg.text in ["Зарегистрироваться как студент", "Зарегистрироваться как преподаватель"])
    def register_user(message):
        """Регистрация пользователя и показ меню в зависимости от роли."""
        role = "student" if message.text == "Зарегистрироваться как студент" else "teacher"
        user_id = message.chat.id
        username = message.from_user.username

        # Добавление пользователя в базу данных
        add_user_to_excel(user_id, username, role)
        bot.send_message(user_id, f"Вы успешно зарегистрировались как {role}!")

        # После регистрации показываем меню
        show_menu(bot, message, role)

    @bot.message_handler(commands=['check_messages'])
    def check_messages_handler(message):
        """Обработчик для проверки новых сообщений."""
        check_for_new_messages(bot, message)

    @bot.message_handler(func=lambda message: True)
    def handle_user_actions(message):
        """Обработка различных действий пользователя."""
        user_id = message.chat.id
        role = get_user_role(user_id)

        if not role:
            bot.send_message(user_id, "Вы не зарегистрированы. Пожалуйста, выполните регистрацию с помощью команды /start.")
            return

        if message.text == "Работа с проектом" and role == "student":
            project_work_menu(bot, message)
        elif message.text == "Связаться с преподавателем" and role == "student":
            bot.send_message(user_id, "Свяжитесь с преподавателем через личные сообщения.")
        elif message.text == "Предложить тему проекта" and role == "student":
            propose_project_topic(bot, message)
        elif message.text == "Мои проекты" and role == "student":
            my_projects_command(bot, message)
        elif message.text == "Просмотреть проекты" and role == "teacher":
            view_projects_command(bot, message)
        elif message.text == "Добавить проект" and role == "teacher":
            bot.send_message(user_id, "Введите название проекта:")
            bot.register_next_step_handler(message, get_project_title)
        else:
            bot.send_message(user_id, "Действие недоступно или команда не распознана.")

    def view_projects_command(bot, message):
        """Просмотр всех проектов преподавателя."""
        projects = get_all_projects()

        if not projects:
            bot.send_message(message.chat.id, "Проекты отсутствуют.")
        else:
            for project in projects:
                bot.send_message(
                    message.chat.id,
                    f"Студент: {project['student_username']} (ID: {project['student_id']})\n"
                    f"Название: {project['title']}\n"
                    f"Описание: {project['description']}\n"
                    f"Статус: {project['status']}\n"
                    f"Назначенный преподаватель: {project['teacher_id'] or 'Не назначен'}\n"
                )

    def get_project_title(message):
        """Получение названия проекта."""
        title = message.text
        bot.send_message(message.chat.id, "Введите описание проекта:")
        bot.register_next_step_handler(message, lambda msg: save_project(msg, title))

    @bot.message_handler(commands=["create_project"])
    def create_project(message):
        """Команда для создания нового проекта."""
        bot.send_message(message.chat.id, "Введите название проекта:")
        bot.register_next_step_handler(message, get_project_title)

    def get_project_title(message):
        """Получение названия проекта и сохранение."""
        title = message.text
        bot.send_message(message.chat.id, "Введите описание проекта:")
        bot.register_next_step_handler(message, lambda msg: save_project(msg, title))

    def save_project(message, title):
        """Сохранение проекта в базе данных."""
        description = message.text
        teacher_id = TEACHER_ID  # Замените на реальный ID преподавателя
        add_project_to_excel(
            student_id=message.chat.id,
            username=message.from_user.username,
            teacher_id=teacher_id,
            title=title,
            description=description,
            status="pending"
        )
        bot.send_message(message.chat.id, "Проект успешно добавлен в базу!")

    def check_for_new_messages(bot, message):
        """Проверка наличия новых сообщений для пользователя."""
        user_id = message.chat.id
        new_messages = get_new_messages_for_user(user_id)

        if new_messages:
            for msg in new_messages:
                bot.send_message(user_id, f"Новое сообщение от {msg['sender_username']}:\n{msg['message_text']}")
                mark_message_as_read(msg['message_id'])
        else:
            bot.send_message(user_id, "Нет новых сообщений.")

    bot.polling(none_stop=True)

def get_user_role(user_id):
    """Функция для получения роли пользователя по его Telegram ID."""
    try:
        workbook = load_workbook(FILE_NAME)
        sheet = workbook["Users"]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:  # Ищем пользователя по ID
                return row[2]  # Возвращаем роль (например, 'student' или 'teacher')

        return None  # Если пользователя нет в базе, возвращаем None
    except Exception as e:
        print(f"Ошибка при получении роли пользователя: {e}")
        return None