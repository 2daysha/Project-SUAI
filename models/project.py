import openpyxl

# Функция для получения всех проектов преподавателя
def get_projects_by_teacher(user_id):
    try:
        workbook = openpyxl.load_workbook('projects.xlsx')
        sheet = workbook.active
        projects = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == user_id:  # проверяем, принадлежит ли проект преподавателю
                projects.append({
                    'title': row[0],
                    'description': row[2],
                    'status': row[3]
                })
        return projects
    except Exception as e:
        print(f"Error getting projects: {e}")
        return []

# Функция для добавления проекта в Excel
def add_project_to_excel(title, teacher_id, description):
    try:
        workbook = openpyxl.load_workbook('projects.xlsx')
        sheet = workbook.active
        sheet.append([title, teacher_id, description, "Не начат"])  # Статус по умолчанию
        workbook.save('projects.xlsx')
    except Exception as e:
        print(f"Error adding project to Excel: {e}")

# Функция для изменения статуса проекта
def change_project_status(project_id, new_status):
    try:
        workbook = openpyxl.load_workbook('projects.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            if row[0] == project_id:
                row[3].value = new_status  # Обновляем статус
                workbook.save('projects.xlsx')
                return True
        return False
    except Exception as e:
        print(f"Error changing project status: {e}")
        return False

# Функция для утверждения проекта (например, установка оценки)
def approve_project(project_id, grade):
    try:
        workbook = openpyxl.load_workbook('projects.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            if row[0] == project_id:
                row[3].value = grade  # Устанавливаем оценку как статус
                workbook.save('projects.xlsx')
                return True
        return False
    except Exception as e:
        print(f"Error approving project: {e}")
        return False
