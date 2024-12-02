import openpyxl

# Функция для получения сообщений от студентов
def get_messages_from_students(teacher_id):
    try:
        workbook = openpyxl.load_workbook('messages.xlsx')
        sheet = workbook.active
        messages = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == teacher_id:
                messages.append({
                    'sender': row[0],
                    'message': row[2]
                })
        return messages
    except Exception as e:
        print(f"Error getting messages: {e}")
        return []

# Функция для отправки сообщения студенту
def send_message_to_student(teacher_id, student_id, message):
    try:
        workbook = openpyxl.load_workbook('messages.xlsx')
        sheet = workbook.active
        sheet.append([student_id, teacher_id, message])
        workbook.save('messages.xlsx')
        return True
    except Exception as e:
        print(f"Error sending message: {e}")
        return False
