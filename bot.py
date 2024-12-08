import telebot
from telebot import types
import openpyxl
import os

TOKEN = '7701145484:AAGc9fPzzhuhwNW8MQebVXc8SUZJYQtg7es'
bot = telebot.TeleBot(TOKEN)

USER_STATE = {}

STUDENT_FILE = 'students.xlsx'
TEACHER_FILE = 'teachers.xlsx'
PROJECTS_FILE = 'projects.xlsx'
PROPOSED_PROJECTS_FILE = 'proposed_projects.xlsx'


def create_file_if_missing(file, headers):
    try:
        openpyxl.load_workbook(file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file)


create_file_if_missing(STUDENT_FILE, ["ID", "Фамилия", "Имя", "Группа"])
create_file_if_missing(TEACHER_FILE, ["ID", "Фамилия", "Имя", "Предмет"])
create_file_if_missing(PROJECTS_FILE, ["ID Преподавателя", "Название проекта", "Фамилия студента",
                                       "Имя студента", "Группа студента", "Статус", "Оценка"])
create_file_if_missing(PROPOSED_PROJECTS_FILE, ["ID Студента", "Название проекта", "Фамилия студента",
                                                "Имя студента", "Группа студента", "ID Преподавателя"])


def is_user_registered(user_id):
    return is_user_in_file(user_id, STUDENT_FILE) or is_user_in_file(user_id, TEACHER_FILE)


def is_teacher(user_id):
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return True
        return False
    except Exception as e:
        print(f"Ошибка проверки роли преподавателя: {e}")
        return False


def is_student(user_id):
    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return True
        return False
    except Exception as e:
        print(f"Ошибка проверки роли студента: {e}")
        return False


def is_user_in_file(user_id, file):
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(user_id):
                return True
        return False
    except FileNotFoundError:
        return False


@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id
    if is_user_registered(user_id):
        send_main_menu(user_id)
    else:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Зарегистрироваться как студент", callback_data="register_student"))
        markup.add(types.InlineKeyboardButton("Зарегистрироваться как преподаватель", callback_data="register_teacher"))
        bot.send_message(user_id, "Привет! Я помогу тебе организовать работу с проектами. Выберите роль:",
                         reply_markup=markup)


def send_main_menu(user_id):
    role = get_user_role(user_id)
    if role == 'student':
        send_student_menu(user_id)
    elif role == 'teacher':
        send_teacher_menu(user_id)


def get_user_role(user_id):
    if is_user_in_file(user_id, STUDENT_FILE):
        return 'student'
    elif is_user_in_file(user_id, TEACHER_FILE):
        return 'teacher'
    return None


@bot.callback_query_handler(func=lambda call: call.data == "student_menu")
def send_student_menu(user_id):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Связаться с преподавателем", callback_data="contact_teacher"))
    markup.add(types.InlineKeyboardButton("Предложить тему проекта", callback_data="suggest_project"))
    markup.add(types.InlineKeyboardButton("Мои проекты", callback_data="my_projects"))
    # Добавляем кнопку "Вернуться в меню" внизу
    markup.add(types.InlineKeyboardButton("Вернуться в меню", callback_data="back_to_main_menu"))
    bot.send_message(user_id, "Добро пожаловать, студент! Выберите действие:", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "teacher_menu")
def send_teacher_menu(user_id):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Добавить проект", callback_data="add_project"))
    markup.add(types.InlineKeyboardButton("Поиск проекта", callback_data="search_project"))
    markup.add(types.InlineKeyboardButton("Скачать отчет", callback_data="download_report"))
    markup.add(types.InlineKeyboardButton("Изменить статус проекта", callback_data="change_status"))
    markup.add(types.InlineKeyboardButton("Оценить проект", callback_data="evaluate_project"))
    # Добавляем кнопку "Вернуться в меню" внизу
    markup.add(types.InlineKeyboardButton("Вернуться в меню", callback_data="back_to_main_menu"))
    bot.send_message(user_id, "Добро пожаловать, преподаватель! Выберите действие:", reply_markup=markup)

# Обработчик для кнопки "Вернуться в меню"
@bot.callback_query_handler(func=lambda call: call.data == "back_to_main_menu")
def back_to_main_menu(call):
    user_id = call.from_user.id
    send_main_menu(user_id)


@bot.callback_query_handler(func=lambda call: call.data in ["register_student", "register_teacher"])
def register_user(call):
    user_id = call.from_user.id
    USER_STATE[user_id] = {'role': call.data.split("_")[1]}
    bot.send_message(user_id, "Введите вашу фамилию:")
    bot.register_next_step_handler(call.message, get_last_name)


def get_last_name(message):
    user_id = message.from_user.id
    USER_STATE[user_id]['last_name'] = message.text.strip()
    bot.send_message(user_id, "Введите ваше имя:")
    bot.register_next_step_handler(message, get_first_name)


def get_first_name(message):
    user_id = message.from_user.id
    USER_STATE[user_id]['first_name'] = message.text.strip()

    role = USER_STATE[user_id]['role']
    if role == "student":
        bot.send_message(user_id, "Введите вашу группу:")
        bot.register_next_step_handler(message, finalize_student_registration)
    elif role == "teacher":
        bot.send_message(user_id, "Введите ваш предмет:")
        bot.register_next_step_handler(message, finalize_teacher_registration)


def finalize_student_registration(message):
    user_id = message.from_user.id
    group = message.text.strip()

    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        ws.append([user_id, USER_STATE[user_id]['last_name'], USER_STATE[user_id]['first_name'], group])
        wb.save(STUDENT_FILE)
        bot.send_message(user_id, "Регистрация завершена! Вы зарегистрированы как студент.")
        send_student_menu(user_id)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при сохранении данных: {e}")
    finally:
        USER_STATE.pop(user_id, None)


def finalize_teacher_registration(message):
    user_id = message.from_user.id
    subject = message.text.strip()

    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        ws.append([user_id, USER_STATE[user_id]['last_name'], USER_STATE[user_id]['first_name'], subject])
        wb.save(TEACHER_FILE)
        bot.send_message(user_id, "Регистрация завершена! Вы зарегистрированы как преподаватель.")
        send_teacher_menu(user_id)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при сохранении данных: {e}")
    finally:
        USER_STATE.pop(user_id, None)


def initialize_projects_file():
    if not os.path.exists(PROJECTS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(["Project ID", "Title", "Description", "Teacher ID", "Student ID", "Status"])
        wb.save(PROJECTS_FILE)

def get_student_data(student_id):
    """Функция для получения данных студента по его ID"""
    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(student_id):  # Проверка ID как строки для точного сравнения
                return row[1], row[2], row[3]  # Возвращаем фамилию, имя и группу
        return None  # Если студента не нашли
    except Exception as e:
        print(f"Ошибка при получении данных студента: {e}")
        return None

@bot.callback_query_handler(func=lambda call: call.data == "add_project")
def add_project_handler(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return

    bot.send_message(user_id, "Введите название проекта:")
    bot.register_next_step_handler(call.message, get_project_title)


def get_project_title(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    bot.send_message(user_id, "Введите описание проекта:")
    bot.register_next_step_handler(message, get_project_description, project_title)


def get_project_description(message, project_title):
    user_id = message.from_user.id
    project_description = message.text.strip()

    bot.send_message(user_id, "Введите ID первого студента:")
    bot.register_next_step_handler(message, get_student_id, project_title, project_description)


def get_student_id(message, project_title, project_description):
    user_id = message.from_user.id
    student_id = message.text.strip()

    student_data = get_student_data(student_id)
    if student_data:
        student_lastname, student_firstname, student_group = student_data
    else:
        student_lastname = student_firstname = student_group = "Неизвестно"
        bot.send_message(user_id, "Студент с таким ID не найден. Данные будут добавлены как 'Неизвестно'.")

    bot.send_message(user_id, f"Добавлен студент: {student_lastname} {student_firstname}, группа: {student_group}.")
    bot.send_message(user_id, "Введите ID следующего студента или введите 'стоп' для завершения добавления участников.")

    bot.register_next_step_handler(message, add_next_student, project_title, project_description,
                                   [(student_lastname, student_firstname, student_group)], user_id)


def add_next_student(message, project_title, project_description, students_data, user_id):
    student_id = message.text.strip()

    if student_id.lower() == "стоп":
        bot.send_message(user_id, "Завершение добавления участников.")
        bot.send_message(user_id, "Введите статус проекта (например, 'Активен'):")
        bot.register_next_step_handler(message, get_project_status, project_title, project_description, students_data, user_id)
        return

    student_data = get_student_data(student_id)
    if student_data:
        student_lastname, student_firstname, student_group = student_data
        students_data.append((student_lastname, student_firstname, student_group))
        bot.send_message(user_id, f"Добавлен студент: {student_lastname} {student_firstname}, группа: {student_group}.")
    else:
        bot.send_message(user_id, "Студент с таким ID не найден. Попробуйте снова.")

    bot.send_message(user_id, "Введите ID следующего студента или введите 'стоп' для завершения добавления участников.")
    bot.register_next_step_handler(message, add_next_student, project_title, project_description, students_data, user_id)



def get_teacher_id(message, project_title, project_description, students_data):
    user_id = message.from_user.id
    teacher_id = message.text.strip()

    bot.send_message(user_id, "Введите статус проекта (например, 'Активен'):")
    bot.register_next_step_handler(message, get_project_status, project_title, project_description, students_data,
                                   teacher_id)


def get_project_status(message, project_title, project_description, students_data, teacher_id):
    user_id = message.from_user.id
    project_status = message.text.strip()

    bot.send_message(user_id, "Введите оценку проекта (например, 'Не оценено'):")
    bot.register_next_step_handler(message, finalize_project_addition, project_title, project_description, students_data, teacher_id, project_status)



def get_teacher_name(teacher_id):
    """Функция для получения ФИО преподавателя по его ID"""
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(teacher_id):  # Проверка ID как строки для точного сравнения
                return f"{row[1]} {row[2]}"  # Возвращаем Фамилию и Имя преподавателя
        return "Неизвестно"  # Если преподаватель не найден
    except Exception as e:
        print(f"Ошибка при получении данных преподавателя: {e}")
        return "Неизвестно"


def finalize_project_addition(message, project_title, project_description, students_data, teacher_id, project_status):
    user_id = message.from_user.id
    project_evaluation = message.text.strip()

    # Получаем ФИО преподавателя по текущему ID
    teacher_name = get_teacher_name(user_id)

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        # Проверяем наличие заголовков, если они отсутствуют - добавляем
        if ws.max_row == 1:
            ws.append(["ФИО Преподавателя", "Название проекта", "Фамилия студента", "Имя студента", "Группа студента",
                       "Статус", "Оценка"])

        for student in students_data:
            student_lastname, student_firstname, student_group = student
            # Добавляем данные проекта в таблицу
            ws.append([teacher_name, project_title, student_lastname, student_firstname, student_group, project_status,
                       project_evaluation])

        wb.save(PROJECTS_FILE)

        bot.send_message(user_id,
                         f"Проект '{project_title}' успешно добавлен с участниками: {', '.join([f'{s[0]} {s[1]}' for s in students_data])}.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при добавлении проекта: {e}")


@bot.callback_query_handler(func=lambda call: call.data == "search_project")
def search_project_handler(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    bot.send_message(user_id, "Введите ID, название или статус проекта для поиска:")
    bot.register_next_step_handler(call.message, search_project)


def search_project(message):
    user_id = message.from_user.id
    search_query = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        found_projects = {}

        # Проходим по всем строкам в файле
        for row in ws.iter_rows(min_row=2, values_only=True):
            project_id, project_name, student_lastname, student_firstname, student_group, project_status, evaluation = row

            # Если совпадает по ID, названию или статусу
            if (str(search_query) in str(project_id)) or (search_query.lower() in str(project_name).lower()) or (
                    search_query.lower() in str(project_status).lower()):
                # Строка проекта с участниками
                participants = found_projects.get(project_id, [])
                participants.append(f"{student_lastname} {student_firstname} ({student_group})")
                found_projects[project_id] = participants

        if found_projects:
            result_message = "Найденные проекты:\n"
            for project_id, participants in found_projects.items():
                # Получаем проект из файла по ID
                wb = openpyxl.load_workbook(PROJECTS_FILE)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == project_id:
                        project_name = row[1]
                        project_status = row[5]
                        evaluation = row[6]
                        break

                # Добавляем информацию о проекте и участниках
                result_message += f"Название: {project_name}, Участники: {', '.join(set(participants))}, Статус: {project_status}, Оценка: {evaluation}\n"

            bot.send_message(user_id, result_message)
        else:
            bot.send_message(user_id, "Проекты по вашему запросу не найдены.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проектов: {e}")



def get_student_group(user_id):
    try:
        wb = openpyxl.load_workbook("students.xlsx")
        ws = wb.active

        # Ищем строку, в которой находится student_id
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:  # Предполагаем, что student_id находится в первой колонке
                return row[2]  # Предполагаем, что группа находится в третьей колонке

        return "Группа не указана"  # Если не нашли, возвращаем дефолтное значение
    except Exception as e:
        print(f"Ошибка при загрузке группы: {e}")
        return "Группа не указана"


@bot.callback_query_handler(func=lambda call: call.data == "my_projects")
def my_projects(call):
    user_id = call.from_user.id  # ID текущего пользователя
    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        student_projects = []

        # Получаем Фамилию и Имя студента, которые зарегистрированы в students.xlsx
        student_lastname = None
        student_firstname = None
        # Прочитаем информацию о студенте, его фамилию и имя, используя ID
        try:
            student_data = get_student_data(user_id)  # Эта функция должна вернуть фамилию и имя студента
            student_lastname, student_firstname, student_group = student_data
        except:
            bot.send_message(user_id, "Не удалось найти вашу информацию в базе данных.")
            return

        # Если студент найден, то ищем все проекты этого студента
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Сравниваем фамилию и имя студента в проекте с фамилией и именем пользователя
            if row[2].strip().lower() == student_lastname.lower() and row[3].strip().lower() == student_firstname.lower():
                project_info = (f"Название: {row[1]}, "
                                f"Преподаватель: {row[0]}, "
                                f"Группа студента: {row[4]}, "
                                f"Статус: {row[5]}, "
                                f"Оценка: {row[6]}")
                student_projects.append(project_info)

        # Если найдены проекты, отправляем их пользователю
        if student_projects:
            bot.send_message(user_id, "Ваши проекты:\n\n" + "\n\n".join(student_projects))
        else:
            bot.send_message(user_id, "У вас нет проектов.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при загрузке проектов: {e}")

def get_student_data(user_id):
    """Функция для получения данных студента (фамилия, имя, группа) по его ID"""
    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return row[1], row[2], row[3]  # Фамилия, Имя, Группа
        return None
    except Exception as e:
        print(f"Ошибка при получении данных студента: {e}")
        return None


@bot.callback_query_handler(func=lambda call: call.data == "change_status")
def change_status(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    bot.send_message(user_id, "Введите название проекта, для которого хотите изменить статус:")
    bot.register_next_step_handler(call.message, get_project_by_title_for_status)


def get_project_by_title_for_status(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        # Ищем проекты по названию
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1].lower() == project_title.lower():  # Название проекта вторая колонка
                projects_found.append(row)

        if not projects_found:
            bot.send_message(user_id, "Проект с таким названием не найден. Попробуйте снова.")
            return

        if len(projects_found) == 1:
            # Если найден один проект, сразу переходим к изменению статуса
            bot.send_message(user_id, "Введите новый статус проекта (например, 'Завершён', 'Активен'):")
            bot.register_next_step_handler(message, set_project_status, projects_found[0])
        else:
            # Если найдено несколько проектов, уточняем
            markup = types.InlineKeyboardMarkup()
            for idx, project in enumerate(projects_found):
                student_info = f"{project[2]} {project[3]} ({project[4]})"
                markup.add(types.InlineKeyboardButton(f"{project[1]} - {student_info}", callback_data=f"choose_{idx}"))
            bot.send_message(user_id, "Найдено несколько проектов с таким названием. Выберите проект:", reply_markup=markup)
            # Сохраняем найденные проекты для дальнейшего использования
            USER_STATE[user_id] = {'projects_found': projects_found}
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проекта: {e}")

def set_project_status(message, project):
    user_id = message.from_user.id
    new_status = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        # Обновляем статус проекта
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value.lower() == project[1].lower() and row[2].value == project[2]:
                row[5].value = new_status
                wb.save(PROJECTS_FILE)
                bot.send_message(user_id, f"Статус проекта '{project[1]}' успешно обновлён на '{new_status}'.")
                return
        bot.send_message(user_id, "Проект не найден. Попробуйте снова.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при обновлении статуса проекта: {e}")



@bot.callback_query_handler(func=lambda call: call.data == "evaluate_project")
def evaluate_project(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    bot.send_message(user_id, "Введите название проекта, который хотите оценить:")
    bot.register_next_step_handler(call.message, get_project_by_title_for_evaluation)


def get_project_by_title_for_evaluation(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        # Ищем проекты по названию
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1].lower() == project_title.lower():  # Название проекта во второй колонке
                projects_found.append(row)

        if not projects_found:
            bot.send_message(user_id, "Проект с таким названием не найден. Попробуйте снова.")
            return

        if len(projects_found) == 1:
            # Если найден один проект, сразу переходим к оценке
            bot.send_message(user_id, "Введите оценку для проекта:")
            bot.register_next_step_handler(message, set_project_evaluation, projects_found[0])
        else:
            # Если найдено несколько проектов, уточняем
            markup = types.InlineKeyboardMarkup()
            for idx, project in enumerate(projects_found):
                student_info = f"{project[2]} {project[3]} ({project[4]})"
                markup.add(types.InlineKeyboardButton(f"{project[1]} - {student_info}", callback_data=f"evaluate_{idx}"))
            bot.send_message(user_id, "Найдено несколько проектов с таким названием. Выберите проект для оценки:", reply_markup=markup)
            # Сохраняем найденные проекты для дальнейшего использования
            USER_STATE[user_id] = {'projects_found': projects_found}
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проекта: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("evaluate_"))
def choose_project_for_evaluation(call):
    user_id = call.from_user.id
    if user_id not in USER_STATE or 'projects_found' not in USER_STATE[user_id]:
        bot.send_message(user_id, "Ошибка! Попробуйте снова.")
        return

    project_index = int(call.data.split("_")[1])
    project = USER_STATE[user_id]['projects_found'][project_index]
    bot.send_message(user_id, "Введите оценку для проекта:")
    bot.register_next_step_handler(call.message, set_project_evaluation, project)
    # Очистка состояния
    USER_STATE.pop(user_id, None)

def set_project_evaluation(message, project):
    user_id = message.from_user.id
    evaluation = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        # Обновляем оценку проекта
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value.lower() == project[1].lower() and row[2].value == project[2]:
                row[6].value = evaluation  # Обновляем колонку оценки
                wb.save(PROJECTS_FILE)
                bot.send_message(user_id, f"Оценка проекта '{project[1]}' успешно обновлена на '{evaluation}'.")
                return
        bot.send_message(user_id, "Проект не найден. Попробуйте снова.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при обновлении оценки проекта: {e}")


@bot.callback_query_handler(func=lambda call: call.data == "download_report")
def download_report(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return

    try:
        report_file = 'project_report.xlsx'
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        wb.save(report_file)

        with open(report_file, 'rb') as f:
            bot.send_document(user_id, f)
        os.remove(report_file)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при скачивании отчета: {e}")


@bot.callback_query_handler(func=lambda call: call.data == "suggest_project")
def suggest_project(call):
    user_id = call.from_user.id
    bot.send_message(user_id, "Выберите преподавателя, которому хотите предложить тему проекта:")

    markup = types.InlineKeyboardMarkup()
    try:
        wb = openpyxl.load_workbook('teachers.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            teacher_id = row[0]
            teacher_name = f"{row[1]} {row[2]}"
            markup.add(types.InlineKeyboardButton(teacher_name, callback_data=f"teacher_{teacher_id}"))
        bot.send_message(user_id, "Выберите преподавателя из списка:", reply_markup=markup)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при загрузке преподавателей: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("teacher_"))
def teacher_selected(call):
    teacher_id = call.data.split("_")[1]
    user_id = call.from_user.id

    bot.send_message(user_id, "Введите тему проекта, которую вы хотите предложить:")
    bot.register_next_step_handler(call.message, lambda msg: save_project_suggestion(msg, teacher_id))


def save_project_suggestion(message, teacher_id):
    user_id = message.from_user.id
    suggestion = message.text.strip()

    try:
        # Получаем информацию о студенте
        wb_students = openpyxl.load_workbook(STUDENT_FILE)
        ws_students = wb_students.active
        student_info = None
        for row in ws_students.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                student_info = row
                break

        if not student_info:
            bot.send_message(user_id, "Ошибка: ваша регистрация как студента не найдена.")
            return

        student_lastname, student_firstname, student_group = student_info[1:4]

        # Сохраняем предложение
        wb_proposed = openpyxl.load_workbook(PROPOSED_PROJECTS_FILE)
        ws_proposed = wb_proposed.active
        ws_proposed.append([user_id, suggestion, student_lastname, student_firstname, student_group, teacher_id])
        wb_proposed.save(PROPOSED_PROJECTS_FILE)

        bot.send_message(user_id, f"Тема проекта '{suggestion}' успешно предложена преподавателю!")

        # Отправляем сообщение преподавателю
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Одобрить", callback_data=f"approve_{user_id}_{len(suggestion)}"))
        markup.add(types.InlineKeyboardButton("Не одобрить", callback_data=f"reject_{user_id}_{len(suggestion)}"))
        bot.send_message(
            teacher_id,
            f"Студент {student_lastname} {student_firstname}, группа {student_group} предложил тему: '{suggestion}'.\n"
            "Нажмите одну из кнопок ниже для одобрения или отклонения этой темы.",
            reply_markup=markup
        )
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при предложении темы: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("approve_") or call.data.startswith("reject_"))
def approve_or_reject(call):
    bot.answer_callback_query(call.id)
    data = call.data.split("_")
    action = data[0]
    user_id = data[1]
    suggestion_length = int(data[2])  # Длина предложения, переданная в callback_data

    if action == "approve":
        # Логика одобрения проекта
        approve_suggestion(call, user_id, suggestion_length)
    elif action == "reject":
        # Логика отклонения проекта с запросом комментария
        ask_for_rejection_comment(call, user_id, suggestion_length)


def approve_suggestion(call, user_id, suggestion_length):
    try:
        wb = openpyxl.load_workbook(PROPOSED_PROJECTS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == user_id and len(row[1].value) == suggestion_length:
                # Перемещаем проект в основной список
                wb2 = openpyxl.load_workbook(PROJECTS_FILE)
                ws2 = wb2.active
                ws2.append([row[5].value, row[1].value, row[2].value, row[3].value, row[4].value, "Одобрено", "Не оценено"])
                wb2.save(PROJECTS_FILE)

                # Удаляем предложение
                ws.delete_rows(row[0].row)
                wb.save(PROPOSED_PROJECTS_FILE)

                # Уведомляем студента
                bot.send_message(user_id, f"Ваш проект '{row[1].value}' был одобрен преподавателем!")
                bot.send_message(call.from_user.id, f"Проект '{row[1].value}' успешно одобрен.")
                break
    except Exception as e:
        bot.send_message(call.from_user.id, f"Ошибка при одобрении проекта: {e}")


def ask_for_rejection_comment(call, user_id, suggestion_length):
    # Сохраняем данные в состояние пользователя
    USER_STATE[call.from_user.id] = {
        "action": "reject",
        "student_id": user_id,
        "suggestion_length": suggestion_length
    }
    bot.send_message(call.from_user.id, "Введите комментарий для отклонения проекта:")
    bot.register_next_step_handler(call.message, process_rejection_comment)


def process_rejection_comment(message):
    teacher_id = message.from_user.id
    comment = message.text.strip()

    # Получаем сохраненные данные
    state = USER_STATE.get(teacher_id)
    if not state:
        bot.send_message(teacher_id, "Произошла ошибка. Попробуйте снова.")
        return

    user_id = state["student_id"]
    suggestion_length = state["suggestion_length"]

    try:
        # Ищем предложение в файле и удаляем
        wb = openpyxl.load_workbook(PROPOSED_PROJECTS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == user_id and len(row[1].value) == suggestion_length:
                # Уведомляем студента об отклонении и добавляем комментарий
                bot.send_message(user_id, f"Ваш проект '{row[1].value}' был отклонен преподавателем.\nКомментарий: {comment}")
                ws.delete_rows(row[0].row)
                wb.save(PROPOSED_PROJECTS_FILE)
                break

        # Уведомляем преподавателя о завершении операции
        bot.send_message(teacher_id, "Проект успешно отклонен с комментарием.")
    except Exception as e:
        bot.send_message(teacher_id, f"Ошибка при отклонении проекта: {e}")
    finally:
        USER_STATE.pop(teacher_id, None)


@bot.callback_query_handler(func=lambda call: call.data == "contact_teacher")
def contact_teacher(call):
    user_id = call.from_user.id

    # Проверяем, что пользователь - студент
    if not is_student(user_id):
        bot.send_message(user_id, "Ошибка! Вы не студент.")
        return

    # Создаем клавиатуру с преподавателями
    markup = types.InlineKeyboardMarkup()
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            teacher_id = row[0]
            teacher_name = f"{row[1]} {row[2]}"  # Фамилия и Имя преподавателя
            markup.add(types.InlineKeyboardButton(teacher_name, callback_data=f"msg_teacher_{teacher_id}"))

        bot.send_message(user_id, "Выберите преподавателя, с которым хотите связаться:", reply_markup=markup)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при загрузке преподавателей: {e}")


# Обработчик для выбора преподавателя
@bot.callback_query_handler(func=lambda call: call.data.startswith("msg_teacher_"))
def teacher_selected(call):
    teacher_id = call.data.split("_")[2]
    user_id = call.from_user.id

    # Просим студента ввести сообщение для преподавателя
    bot.send_message(user_id, "Введите ваше сообщение для преподавателя:")
    bot.register_next_step_handler(call.message, send_message_to_teacher, teacher_id)


# Функция для отправки сообщения преподавателю
def send_message_to_teacher(message, teacher_id):
    user_id = message.from_user.id
    user_message = message.text.strip()

    # Получаем данные студента из базы
    student_data = get_student_data(user_id)
    if student_data:
        student_lastname, student_firstname, student_group = student_data
    else:
        student_lastname = message.from_user.last_name or "Неизвестно"
        student_firstname = message.from_user.first_name or "Неизвестно"
        student_group = "Неизвестно"

    try:
        # Проверяем, что преподаватель существует
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        teacher_found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == teacher_id:  # Если ID преподавателя совпадает
                teacher_found = True
                teacher_name = f"{row[1]} {row[2]}"
                break

        if teacher_found:
            # Отправляем сообщение преподавателю с кнопкой "Ответить"
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Ответить студенту", callback_data=f"reply_{user_id}"))
            bot.send_message(
                teacher_id,
                f"Сообщение от студента {student_lastname} {student_firstname}, группа {student_group}:\n{user_message}",
                reply_markup=markup
            )
            bot.send_message(user_id, f"Ваше сообщение было отправлено преподавателю {teacher_name}.")
        else:
            bot.send_message(user_id, "Преподаватель не найден.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при отправке сообщения: {e}")


# Обработчик для ответа преподавателя
@bot.callback_query_handler(func=lambda call: call.data.startswith("reply_"))
def reply_to_student(call):
    teacher_id = call.from_user.id
    student_id = call.data.split("_")[1]

    # Проверяем, что преподаватель существует
    if not is_teacher(teacher_id):
        bot.send_message(teacher_id, "Ошибка! Вы не преподаватель.")
        return

    bot.send_message(teacher_id, "Введите ваше сообщение для студента:")
    bot.register_next_step_handler(call.message, send_reply_to_student, student_id)


# Функция для отправки ответа студенту
def send_reply_to_student(message, student_id):
    teacher_id = message.from_user.id
    teacher_message = message.text.strip()

    # Получаем данные преподавателя
    teacher_data = get_teacher_data(teacher_id)
    if teacher_data:
        teacher_lastname, teacher_firstname = teacher_data
    else:
        teacher_lastname = teacher_firstname = "Неизвестно"

    try:
        # Отправляем сообщение студенту
        bot.send_message(
            student_id,
            f"Ответ от преподавателя {teacher_lastname} {teacher_firstname}:\n{teacher_message}"
        )
        bot.send_message(teacher_id, "Ваше сообщение было отправлено студенту.")
    except Exception as e:
        bot.send_message(teacher_id, f"Ошибка при отправке сообщения студенту: {e}")


# Вспомогательная функция для получения данных преподавателя
def get_teacher_data(teacher_id):
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(teacher_id):  # Проверяем ID преподавателя
                return row[1], row[2]  # Возвращаем Фамилию и Имя преподавателя
        return None
    except Exception as e:
        print(f"Ошибка при получении данных преподавателя: {e}")
        return None


# Запуск бота
if __name__ == "__main__":
    bot.polling(none_stop=True)