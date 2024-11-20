from telebot import types
from excel_manager import get_projects_by_teacher, add_message_to_excel
from openpyxl import load_workbook

def teacher_menu(bot, message):
    """
    Главное меню преподавателя.
    """
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("Просмотреть проекты")
    btn2 = types.KeyboardButton("Добавить проект")
    btn3 = types.KeyboardButton("Отправить сообщение студенту")
    markup.add(btn1, btn2, btn3)

    bot.send_message(message.chat.id, "Меню преподавателя. Выберите действие:", reply_markup=markup)

def send_message_to_student(bot, message):
    """
    Запрашивает выбор проекта для отправки сообщения студенту.
    """
    try:
        projects = get_projects_by_teacher(message.chat.id)
        if not projects:
            bot.send_message(message.chat.id, "У вас нет проектов.")
            return

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for project in projects:
            markup.add(types.KeyboardButton(f"Проект: {project['title']}"))

        bot.send_message(
            message.chat.id,
            "Выберите проект, чтобы отправить сообщение студенту:",
            reply_markup=markup
        )
        bot.register_next_step_handler(message, handle_message_to_student, projects, bot)
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка при получении проектов: {str(e)}")

def handle_message_to_student(message, projects, bot):
    """
    Обработка выбора проекта для отправки сообщения студенту.
    """
    try:
        selected_project = next(
            (project for project in projects if message.text == f"Проект: {project['title']}"),
            None
        )

        if selected_project:
            bot.send_message(message.chat.id, "Введите сообщение для студента:")
            bot.register_next_step_handler(message, send_message_to_student_final, selected_project, bot)
        else:
            bot.send_message(message.chat.id, "Неверный выбор. Попробуйте еще раз.")
            bot.register_next_step_handler(message, handle_message_to_student, projects, bot)
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка обработки выбора: {str(e)}")

def send_message_to_student_final(message, project, bot):
    """
    Финальная отправка сообщения студенту.
    """
    student_id = project.get('student_id')
    if not student_id:
        bot.send_message(message.chat.id, "Ошибка: У выбранного проекта нет назначенного студента.")
        return

    try:
        bot.send_message(student_id, f"Сообщение от преподавателя:\n{message.text}")
        bot.send_message(message.chat.id, "Сообщение успешно отправлено студенту.")

        # Логируем сообщение в базе данных
        add_message_to_excel(
            sender_id=message.chat.id,
            receiver_id=student_id,
            message_text=message.text,
            project_id=project['id']
        )
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка при отправке сообщения: {str(e)}")

def add_message_to_excel(sender_id, receiver_id, message_text, project_id):
    """
    Добавление сообщения в Excel для записи.
    """
    try:
        wb = load_workbook('messages.xlsx')
        sheet = wb.active

        row = sheet.max_row + 1  # Определяем первую свободную строку

        # Добавляем новое сообщение
        sheet[f'A{row}'] = row - 1  # ID сообщения
        sheet[f'B{row}'] = sender_id
        sheet[f'C{row}'] = receiver_id
        sheet[f'D{row}'] = message_text
        sheet[f'E{row}'] = 'unread'  # Статус сообщения
        sheet[f'F{row}'] = project_id  # ID проекта

        wb.save('messages.xlsx')
    except FileNotFoundError:
        raise FileNotFoundError("Файл 'messages.xlsx' не найден. Проверьте его расположение.")
    except Exception as e:
        raise Exception(f"Ошибка при добавлении сообщения в Excel: {str(e)}")
